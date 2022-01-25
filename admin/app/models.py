# -*- coding: utf-8 -*-
"""
aid.models
~~~~~~~~~~~~~~~~~~~

访问数据库
"""

from openpyxl import load_workbook
from pathlib import Path
from pyape.app import gdb
from pyape.app import logger


class Prize(gdb.Model):
    """ 奖品表
    """
    __tablename__ = 'prize'

    # 礼物唯一 ID
    prize_id = gdb.Column(gdb.INT, primary_key=True, nullable=False, autoincrement=True)

    # 奖项级别，1 代表一等奖
    level = gdb.Column(gdb.INT, nullable=False)

    # 奖项名称
    level_name = gdb.Column(gdb.VARCHAR(50), nullable=False)

    # 奖品名称
    prize_name = gdb.Column(gdb.VARCHAR(100), nullable=False)

    # 奖品图像文件名称
    image_name = gdb.Column(gdb.VARCHAR(200), nullable=False)

    def __repr__(self):
        return f'<Prize {self.level_name}:{self.prize_name}({self.prize_id})>'


class Round(gdb.Model):
    """ 抽奖轮
    """
    __tablename__ = 'round'

    # 轮次，一个轮次可以包含多个奖项
    round = gdb.Column(gdb.INT, primary_key=True, nullable=False)

    # 奖品 ID，与轮次对应，一起成为主键
    prize_id = gdb.Column(gdb.INT, primary_key=True, nullable=False)

    # 中奖者 user_id，默认为 null
    user_id = gdb.Column(gdb.INT, nullable=True)

    def __repr__(self):
        return '<Round %r.%r>' % (self.round, self.prize_id)


class User(gdb.Model):
    """ 用户表
    """
    __tablename__ = 'user'

    # 用户 id
    user_id = gdb.Column(gdb.INT, primary_key=True, nullable=False, autoincrement=True)

    # 用户姓名
    name = gdb.Column(gdb.VARCHAR(50), nullable=False)

    # 用户所属部门
    department = gdb.Column(gdb.VARCHAR(100), nullable=False)

    def __repr__(self):
        return '<User %r.%r>' % (self.user_id, self.name)


def init_from_xlsx() -> None:
    """ 从 data 文件夹中的 xlsx 获取初始数据并写入数据库
    """
    from pyape import gconfig
    data_files = {'prize.xlsx': Prize,
        'user.xlsx': User, 
        'round.xlsx': Round
    }
    for f in data_files.keys():
        file: Path = gconfig.getdir(f'data/{f}')
        logger.info(f'载入 Excel 文件 {file}')
        wb = load_workbook(file, read_only=True, data_only=True)
        rows = []
        for row in wb.active.iter_rows():
            # 跳过标题行
            if row[0].row == 1:
                continue
            db_row = None
            if f.startswith('prize'):
                db_row = Prize(
                    prize_id=row[0].value, 
                    level=row[1].value,
                    level_name=row[2].value,
                    prize_name=row[3].value,
                    image_name=row[4].value)
            elif f.startswith('user'):
                db_row = User(
                    user_id=row[0].value,
                    name=row[1].value,
                    department=row[2].value,
                )
            elif f.startswith('round'):
                db_row = Round(
                    round=row[0].value,
                    prize_id=row[1].value,
                )
            rows.append(db_row)
        gdb.session.add_all(rows)
        gdb.session.commit()
        

    